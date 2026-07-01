from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_TEMPLATE_DIR = Path(__file__).resolve().parents[4] / "문서" / "LSS ERP"


PROJECT_HEADERS = [
    ("JOB NO", "project_no"),
    ("프로젝트명", "project_name"),
    ("계약업체명", "client_name"),
    ("발주처", "order_company"),
    ("사업부", "business_division"),
    ("팀", "team_name"),
    ("사업구분", "business_category"),
    ("영업 담당자", "sales_manager"),
    ("실행 담당자", "execution_manager"),
    ("수금담당자", "collection_manager"),
    ("구분(매출유형)", "revenue_type"),
    ("공종", "work_type"),
    ("수금조건", "collection_terms"),
    ("하자보증기간", "warranty_period"),
    ("원도급/하도급", "contract_form"),
    ("현장 주소지", "site_address"),
    ("내수/해외", "contract_type"),
    ("특수관계", "special_relation"),
    ("진행/종료", "status"),
    ("계약일", "contract_start"),
    ("준공일", "construct_end"),
    ("개월수", "months"),
    ("총 계약금액", "contract_amount"),
    ("자재비", "sales_material_cost_total"),
    ("노무비", "sales_labor_cost"),
]

SALES_HEADERS = [
    ("작성차수", "entry_round"),
    ("영업번호", "sales_no"),
    ("프로젝트명", "project_name"),
    ("발주처", "client_name"),
    ("수주확도", "probability"),
    ("영업상태", "sales_status"),
    ("프로젝트번호", "project_no"),
    ("사업부", "business_division"),
    ("영업팀", "sales_team"),
    ("사업구분", "business_category"),
    ("담당자", "manager"),
    ("구분(매출유형)", "revenue_type"),
    ("계약예정일", "contract_expected_date"),
    ("준공예정일", "completion_expected_date"),
    ("내수/해외", "domestic_overseas"),
    ("특수관계", "special_relation"),
    ("재료비율%", "material_ratio"),
    ("발주예상금액", "expected_order_amount"),
]

MATERIAL_HEADERS = [
    ("회사코드", "material_company_group_code"),
    ("품번", "material_code"),
    ("품명", "material_name"),
    ("규격", "spec"),
    ("재고단위", "unit"),
    ("관리단위", "management_unit"),
    ("환산계수", "conversion_factor"),
    ("계정구분", "material_type"),
    ("조달구분", "procurement_type"),
    ("품목군코드", "item_group_code"),
    ("품목군명", "item_group_name"),
    ("LOT여부", "lot_use_yn"),
    ("검사여부", "inspection_type"),
    ("LOT수량", "lot_quantity"),
    ("도면번호", "drawing_no"),
    ("HS CODE", "hs_code"),
    ("폭", "width_value"),
    ("폭단위", "width_unit"),
    ("높이", "height_value"),
    ("높이단위", "height_unit"),
    ("깊이", "depth_value"),
    ("깊이단위", "depth_unit"),
    ("중량", "weight_value"),
    ("중량단위", "weight_unit"),
    ("면적", "area_value"),
    ("면적단위", "area_unit"),
    ("SET품목", "set_item_yn"),
    ("사용여부", "use_yn"),
    ("BATCH수량", "batch_quantity"),
    ("BARCODE", "barcode"),
    ("재질", "material_quality"),
    ("길이", "length_value"),
    ("길이단위", "length_unit"),
    ("비중", "density_value"),
    ("과세구분", "tax_type"),
    ("WEB주문여부", "web_order_yn"),
    ("표준단가", "standard_price"),
    ("비고", "notes"),
    ("주문비고", "order_notes"),
]

COMPANY_HEADERS = [
    ("회사코드", "company_group_code"),
    ("금융기관코드", "payment_bank_code"),
    ("금융기관", "bank_name"),
    ("거래처코드", "company_code"),
    ("거래처약칭", "short_name"),
    ("거래처명", "company_name"),
    ("거래처구분", "company_type"),
    ("사업자등록번호", "business_no"),
    ("대표자명", "ceo_name"),
    ("업태", "business_type"),
    ("종목", "business_item"),
    ("우편번호", "postal_code"),
    ("주소상세1", "address_detail1"),
    ("주소상세2", "address_detail2"),
    ("전화번호", "phone"),
    ("팩스번호", "fax"),
    ("홈페이지", "homepage"),
    ("메일주소", "email"),
    ("이메일", "email"),
    ("국가코드", "country_code"),
    ("거래처분류코드", "company_category_code"),
    ("거래처분류명", "company_category_name"),
    ("거래처등급코드", "company_grade_code"),
    ("거래처등급명", "company_grade_name"),
    ("지역코드", "region_code"),
    ("지역명", "region_name"),
    ("전자세금계산서여부", "electronic_tax_invoice_yn"),
    ("사용여부", "use_yn"),
    ("거래시작일", "transaction_start_date"),
    ("거래상태", "transaction_status"),
    ("결제조건", "payment_terms"),
    ("은행명", "bank_name"),
    ("계좌번호", "bank_account"),
    ("예금 계좌번호", "bank_account"),
    ("예금주", "bank_holder"),
    ("여신한도", "credit_limit"),
    ("비고", "notes"),
]

COMPANY_TEMPLATE_COLUMNS = [
    ("회사코드", "company_group_code", "CO_CD", "타입 : 문자\n길이 : 4\n필수 : True\n설명 :"),
    ("거래처 코드", "company_code", "TR_CD", "타입 : 문자\n길이 : 10\n필수 : True\n설명 :"),
    ("거래처 구분", "company_type", "TR_FG", "타입 : 문자\n길이 : 1\n필수 : True\n설명 : 1.일반, 2.수출, 3.수입"),
    ("거래처명", "company_name", "TR_NM", "타입 : 문자\n길이 : 60\n필수 : False\n설명 :"),
    ("거래처명 약칭", "short_name", "ATTR_NM", "타입 : 문자\n길이 : 60\n필수 : False\n설명 :"),
    ("사업자등록번호", "business_no", "REG_NB", "타입 : 문자\n길이 : 20\n필수 : False\n설명 :"),
    ("대표자명", "ceo_name", "CEO_NM", "타입 : 문자\n길이 : 30\n필수 : False\n설명 :"),
    ("업태", "business_type", "BUSINESS", "타입 : 문자\n길이 : 40\n필수 : False\n설명 :"),
    ("종목", "business_item", "JONGMOK", "타입 : 문자\n길이 : 40\n필수 : False\n설명 :"),
    ("주소 상세1", "address_detail1", "DIV_ADDR1", "타입 : 문자\n길이 : 60\n필수 : False\n설명 :"),
    ("전화번호", "phone", "TEL", "타입 : 문자\n길이 : 20\n필수 : False\n설명 :"),
    ("팩스번호", "fax", "FAX", "타입 : 문자\n길이 : 20\n필수 : False\n설명 :"),
    ("이메일", "email", "EMAIL", "타입 : 문자\n길이 : 40\n필수 : False\n설명 :"),
    ("금융기관코드", "payment_bank_code", "", ""),
    ("금융기관", "bank_name", "", ""),
    ("예금 계좌번호", "bank_account", "BA_NB", "타입 : 문자\n길이 : 40\n필수 : False\n설명 :"),
    ("예금주", "bank_holder", "", ""),
]

MONTH_LABELS = [f"{month}월" for month in range(1, 13)]

FIELD_ALIASES = {
    "project_no": ["job no", "jobno", "프로젝트번호", "pjt no", "pjt no."],
    "project_name": ["프로젝트명", "프로젝트"],
    "client_name": ["계약업체명", "발주처", "거래처명", "거래처"],
    "business_division": ["사업부"],
    "team_name": ["팀", "pm 부서"],
    "sales_team": ["영업팀"],
    "business_category": ["사업구분"],
    "sales_manager": ["영업 담당자", "영업담당자"],
    "execution_manager": ["실행 담당자", "실행담당자"],
    "collection_manager": ["수금담당자"],
    "manager": ["담당자"],
    "revenue_type": ["구분(매출유형)", "매출유형"],
    "work_type": ["공종"],
    "collection_terms": ["수금조건"],
    "warranty_period": ["하자보증기간"],
    "contract_form": ["원도급/하도급"],
    "site_address": ["현장 주소지", "현장주소지"],
    "contract_type": ["내수/해외"],
    "domestic_overseas": ["내수/해외"],
    "special_relation": ["특수관계"],
    "status": ["진행/종료", "진행상태"],
    "contract_start": ["계약일"],
    "construct_end": ["준공일"],
    "contract_expected_date": ["계약예정일"],
    "completion_expected_date": ["준공예정일"],
    "contract_amount": ["총 계약금액", "계약금액"],
    "material_ratio": ["재료비율%", "재료비율", "자재비율"],
    "expected_order_amount": ["발주예상금액"],
    "entry_round": ["작성차수"],
    "sales_no": ["영업번호"],
    "probability": ["수주확도"],
    "sales_status": ["영업상태"],
    "material_code": ["품번", "자재코드", "품목코드", "코드", "ITEM_CD"],
    "material_company_group_code": ["회사코드", "CO_CD"],
    "material_name": ["품명", "자재명", "품목명", "이름", "명칭", "ITEM_NM"],
    "spec": ["규격", "spec", "ITEM_DC"],
    "material_type": ["계정구분", "ACCT_FG"],
    "procurement_type": ["조달구분", "ODR_FG"],
    "unit": ["입고단위", "재고단위", "단위", "UNIT_DC"],
    "management_unit": ["관리단위", "UNITMANG_DC"],
    "conversion_factor": ["환산계수", "UNITCHNG_NB"],
    "use_yn": ["사용여부", "USE_YN"],
    "lot_use_yn": ["LOT여부", "LOT_FG"],
    "set_item_yn": ["SET여부", "SETITEM_FG"],
    "inspection_type": ["검사여부", "QC_FG"],
    "standard_price": ["표준단가", "단가", "가격"],
    "company_group_code": ["회사코드", "CO_CD"],
    "company_code": ["거래처코드", "거래처 코드", "코드", "TR_CD"],
    "company_type": ["거래처구분", "거래처 구분", "TR_FG"],
    "short_name": ["거래처약칭", "거래처명약칭", "거래처명 약칭", "약칭", "ATTR_NM"],
    "company_name": ["거래처명", "상호", "업체명", "TR_NM"],
    "business_no": ["사업자등록번호", "사업자번호", "REG_NB"],
    "ceo_name": ["대표자명", "대표자", "CEO_NM"],
    "business_type": ["업태", "BUSINESS"],
    "business_item": ["종목", "JONGMOK"],
    "postal_code": ["우편번호"],
    "address_detail1": ["주소상세1", "주소 상세1", "사업장주소", "주소", "DIV_ADDR1"],
    "address_detail2": ["주소상세2", "상세주소"],
    "phone": ["전화번호", "전화", "TEL"],
    "fax": ["팩스번호", "팩스", "FAX"],
    "email": ["메일주소", "이메일", "email", "EMAIL"],
    "payment_bank_code": ["금융기관코드"],
    "bank_name": ["은행명", "금융기관"],
    "bank_account": ["계좌번호", "예금계좌번호", "예금 계좌번호", "BA_NB"],
    "bank_holder": ["예금주", "예금주명"],
}


def canonical_header(value: Any) -> str:
    return "".join(str(value or "").strip().lower().split())


def alias_lookup(fields: Iterable[str]) -> dict[str, str]:
    lookup = {}
    for field in fields:
        lookup[canonical_header(field)] = field
        for alias in FIELD_ALIASES.get(field, []):
            lookup[canonical_header(alias)] = field
    for month in MONTH_LABELS:
        lookup[canonical_header(month)] = month
        lookup[canonical_header(f"{month}매출")] = f"{month}매출"
        lookup[canonical_header(f"{month}수금")] = f"{month}수금"
    return lookup


def read_upload_rows(file: UploadFile, fields: Iterable[str], min_matches: int = 2) -> list[dict[str, Any]]:
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
    raw_rows = read_workbook_rows(content, file.filename or "")
    return normalize_table_rows(raw_rows, fields, min_matches=min_matches)


def read_upload_rows_with_raw_headers(file: UploadFile, fields: Iterable[str], min_matches: int = 2) -> list[dict[str, Any]]:
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
    raw_rows = read_workbook_rows(content, file.filename or "")
    return normalize_table_rows_with_raw_headers(raw_rows, fields, min_matches=min_matches)


def read_workbook_rows(content: bytes, filename: str) -> list[list[Any]]:
    normalized_filename = (filename or "").lower()
    if normalized_filename.endswith((".xlsx", ".xlsm")):
        return _read_xlsx_rows(content)
    elif normalized_filename.endswith(".xls"):
        return _read_xls_rows(content)
    raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xlsm, .xls)만 업로드할 수 있습니다.")


def _read_xlsx_rows(content: bytes) -> list[list[Any]]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]


def _read_xls_rows(content: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="구형 .xls 파일을 읽으려면 xlrd 패키지가 필요합니다.",
        ) from exc
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = []
    for row_idx in range(sheet.nrows):
        rows.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])
    return rows


def normalize_table_rows(raw_rows: list[list[Any]], fields: Iterable[str], min_matches: int = 2) -> list[dict[str, Any]]:
    lookup = alias_lookup(fields)
    best_index = -1
    best_mapping: dict[int, str] = {}
    best_score = 0
    for row_index, row in enumerate(raw_rows[:20]):
        mapping = {}
        duplicate_counts = {}
        for col_index, value in enumerate(row):
            field = lookup.get(canonical_header(value))
            if field:
                mapping[col_index] = resolve_duplicate_header(field, duplicate_counts)
        if len(mapping) > best_score:
            best_score = len(mapping)
            best_index = row_index
            best_mapping = mapping
    if best_index < 0 or best_score < min_matches:
        raise HTTPException(status_code=400, detail="엑셀 헤더를 찾을 수 없습니다. 양식의 헤더 행을 확인해 주세요.")

    rows = []
    for source in raw_rows[best_index + 1:]:
        item = {}
        for col_index, field in best_mapping.items():
            if col_index < len(source):
                value = clean_cell_value(source[col_index])
                if value not in (None, ""):
                    item[field] = value
        if item:
            rows.append(item)
    return rows


def normalize_table_rows_with_raw_headers(raw_rows: list[list[Any]], fields: Iterable[str], min_matches: int = 2) -> list[dict[str, Any]]:
    best_index, best_mapping = detect_header_mapping(raw_rows, fields, min_matches)
    raw_headers = build_raw_header_mapping(raw_rows[best_index])
    rows = []
    for source in raw_rows[best_index + 1:]:
        item = {}
        for col_index, field in best_mapping.items():
            if col_index < len(source):
                value = clean_cell_value(source[col_index])
                if value not in (None, ""):
                    item[field] = value
        raw_item = {}
        for col_index, header in raw_headers.items():
            if col_index < len(source):
                value = clean_cell_value(source[col_index])
                if value not in (None, ""):
                    raw_item[header] = value
        if item or raw_item:
            rows.append({"mapped": item, "raw": raw_item})
    return rows


def detect_header_mapping(raw_rows: list[list[Any]], fields: Iterable[str], min_matches: int = 2) -> tuple[int, dict[int, str]]:
    lookup = alias_lookup(fields)
    best_index = -1
    best_mapping: dict[int, str] = {}
    best_score = 0
    for row_index, row in enumerate(raw_rows[:20]):
        mapping = {}
        duplicate_counts = {}
        for col_index, value in enumerate(row):
            field = lookup.get(canonical_header(value))
            if field:
                mapping[col_index] = resolve_duplicate_header(field, duplicate_counts)
        if len(mapping) > best_score:
            best_score = len(mapping)
            best_index = row_index
            best_mapping = mapping
    if best_index < 0 or best_score < min_matches:
        raise HTTPException(status_code=400, detail="엑셀 헤더를 찾을 수 없습니다. 양식의 헤더 행을 확인해 주세요.")
    return best_index, best_mapping


def build_raw_header_mapping(header_row: list[Any]) -> dict[int, str]:
    mapping = {}
    for col_index, value in enumerate(header_row):
        label = str(value or "").strip()
        if not label:
            continue
        mapping[col_index] = f"{col_index + 1:03d}_{label}"
    return mapping


def resolve_duplicate_header(field: str, duplicate_counts: dict[str, int]) -> str:
    duplicate_counts[field] = duplicate_counts.get(field, 0) + 1
    occurrence = duplicate_counts[field]
    if field in MONTH_LABELS:
        if occurrence == 1:
            return f"order_current_{field}"
        if occurrence == 2:
            return f"order_next_{field}"
        if occurrence == 3:
            return f"input_current_{field}"
        return f"{field}_{occurrence}"
    if field.endswith("매출"):
        month = field.replace("매출", "")
        if occurrence == 1:
            return f"revenue_current_{month}"
        if occurrence == 2:
            return f"revenue_next_{month}"
        return f"revenue_{occurrence}_{month}"
    return field


def clean_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def to_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from xlrd.xldate import xldate_as_datetime

            return xldate_as_datetime(value, 0).date()
        except Exception:
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%y-%m-%d", "%y.%m.%d", "%y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_decimal_value(value: Any, default: Decimal | None = None) -> Decimal | None:
    if value in (None, ""):
        return default
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return default
    try:
        return Decimal(text)
    except InvalidOperation:
        return default


def make_template_response(title: str, headers: list[tuple[str, str]], filename: str) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_index, (label, _) in enumerate(headers, start=1):
        cell = ws.cell(1, col_index, label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(len(label) + 4, 12), 24)
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    return output.getvalue(), filename


def make_source_header_template_response(
    source_filename: str,
    download_filename: str,
    header_rows: int,
    fallback: tuple[bytes, str] | None = None,
) -> tuple[bytes, str]:
    source_path = SOURCE_TEMPLATE_DIR / source_filename
    if not source_path.exists():
        if fallback:
            return fallback
        raise HTTPException(status_code=404, detail=f"원본 엑셀 양식을 찾을 수 없습니다: {source_filename}")

    suffix = source_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        raw_rows = _read_xlsx_rows(source_path.read_bytes())[:header_rows]
    elif suffix == ".xls":
        raw_rows = _read_xls_rows(source_path.read_bytes())[:header_rows]
    else:
        if fallback:
            return fallback
        raise HTTPException(status_code=400, detail="지원하지 않는 원본 엑셀 양식입니다.")

    return make_header_rows_template_response(source_path.stem[:31], raw_rows, download_filename)


def make_header_rows_template_response(title: str, rows: list[list[Any]], filename: str) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Template"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    code_fill = PatternFill("solid", fgColor="FFFACD")
    meta_fill = PatternFill("solid", fgColor="808080")
    max_cols = max((len(row) for row in rows), default=0)

    for row_index, row in enumerate(rows, start=1):
        for col_index in range(1, max_cols + 1):
            value = row[col_index - 1] if col_index - 1 < len(row) else None
            cell = ws.cell(row_index, col_index, clean_cell_value(value))
            cell.alignment = Alignment(
                horizontal="center" if row_index < 3 else "left",
                vertical="center" if row_index < 3 else "top",
                wrap_text=True,
            )
            if row_index == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            elif row_index == 2:
                cell.fill = code_fill
            elif row_index == 3:
                cell.fill = meta_fill
                cell.font = Font(color="FFFFFF", size=9)
    for col_index in range(1, max_cols + 1):
        width_source = next((row[col_index - 1] for row in rows if col_index - 1 < len(row) and row[col_index - 1]), "")
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(len(str(width_source)) + 4, 12), 24)
    if len(rows) >= 3:
        ws.row_dimensions[3].height = 48
    ws.freeze_panes = f"A{len(rows) + 1}"
    output = BytesIO()
    wb.save(output)
    return output.getvalue(), filename


def make_company_template_response(filename: str) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "거래처 관리"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    code_fill = PatternFill("solid", fgColor="FFFACD")
    meta_fill = PatternFill("solid", fgColor="808080")
    for col_index, (label, _field, code, meta) in enumerate(COMPANY_TEMPLATE_COLUMNS, start=1):
        label_cell = ws.cell(1, col_index, label)
        label_cell.font = Font(bold=True)
        label_cell.fill = header_fill
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        code_cell = ws.cell(2, col_index, code)
        code_cell.fill = code_fill
        code_cell.alignment = Alignment(horizontal="center", vertical="center")

        meta_cell = ws.cell(3, col_index, meta)
        meta_cell.fill = meta_fill
        meta_cell.font = Font(color="FFFFFF", size=9)
        meta_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(len(label) + 8, 14), 24)
    ws.row_dimensions[3].height = 48
    ws.freeze_panes = "A4"
    output = BytesIO()
    wb.save(output)
    return output.getvalue(), filename
